"""Tests for the ratings module: Excel parsing, fantavoto, upsert, and the consistency check."""

from __future__ import annotations

import io

from openpyxl import Workbook

from euroleghe_ingest.config import Config
from euroleghe_ingest.db.database import init_db
from euroleghe_ingest.modules import ratings
from euroleghe_ingest.modules.validate import check_ratings_consistency

HEADER = ["Cod.", "Ruolo", "Squadra", "Nome", "Voto", "Gf", "Gs", "Rp", "Rs", "Rf", "Au", "Amm", "Esp", "Ass"]
ROWS = [
    ["Voti EuroLeghe 5a giornata", None, None, None, None, None, None, None, None, None, None, None, None, None],
    ["disclaimer", None, None, None, None, None, None, None, None, None, None, None, None, None],
    ["Bayern", None, None, None, None, None, None, None, None, None, None, None, None, None],
    HEADER,
    [100, "P", "Bayern", "Neuer", 6, 0, 1, 0, 0, 0, 0, 0, 0, 0],            # GK conceded 1 -> fanta 5.0
    # 2 goals + a SCORED penalty (Rf, not Rs - Rs = "rigori sbagliati") + a yellow + an assist
    [200, "A", "Bayern", "Kane", 7, 2, 0, 0, 0, 1, 0, 1, 0, 1],
    [300, "ALL", "Bayern", "Kompany", 6.5, 0, 0, 0, 0, 0, 0, 0, 0, 0],      # coach
    [400, "C", "Bayern", "Benched", None, 0, 0, 0, 0, 0, 0, 0, 0, 0],       # no vote
]


def _xlsx_bytes(rows):
    wb = Workbook()
    ws = wb.active
    ws.title = "Statistico"
    for r in rows:
        ws.append(r)
    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue()


def _scoring():
    return Config().load_scoring()


def test_parse_workbook():
    recs = ratings.parse_workbook(_xlsx_bytes(ROWS), "2023-24", 5)
    assert len(recs) == 4
    by_id = {r["fc_id"]: r for r in recs}
    assert by_id[300]["role"] == "ALL"                      # coach kept
    kane = by_id[200]
    assert kane["team"] == "Bayern"
    assert kane["canon"]["goals"] == 2 and kane["canon"]["assists"] == 1
    assert kane["canon"]["pen_scored"] == 1 and kane["canon"]["yellows"] == 1
    assert set(kane["raw"]) == {"Voto", "Gf", "Gs", "Rp", "Rs", "Rf", "Au", "Amm", "Esp", "Ass"}
    assert by_id[400]["canon"]["mv"] is None                # no vote


def test_parse_empty_sheet_stops():
    # Beyond the season the endpoint returns disclaimer+header but no player rows -> no records.
    empty = ROWS[:4]  # the three text rows + header, no players
    assert ratings.parse_workbook(_xlsx_bytes(empty), "2023-24", 39) == []


def test_compute_fantavoto():
    recs = {r["fc_id"]: r for r in ratings.parse_workbook(_xlsx_bytes(ROWS), "2023-24", 5)}
    s = _scoring()
    assert ratings.compute_fantavoto(recs[100]["canon"], s) == 5.0          # 6 - 1 conceded
    assert ratings.compute_fantavoto(recs[200]["canon"], s) == 16.5         # 7 +6 +1 +3 -0.5
    assert ratings.compute_fantavoto(recs[400]["canon"], s) is None         # no vote
    # GK clean-sheet is intentionally excluded (mirrors the source fantavoto): 6 - 0 = 6.0
    assert ratings.compute_fantavoto({"mv": 6.0, "goals_conceded": 0}, s) == 6.0


def test_upsert(tmp_path):
    conn = init_db(tmp_path / "euro.db")
    recs = ratings.parse_workbook(_xlsx_bytes(ROWS), "2023-24", 5)
    n = ratings.upsert_records(conn, recs, _scoring())
    assert n == 4
    assert conn.execute("SELECT COUNT(*) FROM match_ratings").fetchone()[0] == 4
    # raw bonuses: 4 players x 10 columns
    assert conn.execute("SELECT COUNT(*) FROM match_rating_bonuses").fetchone()[0] == 40
    status = dict(conn.execute("SELECT fc_id, status FROM match_ratings").fetchall())
    assert status[100] == "played" and status[400] == "no_vote"
    assert conn.execute("SELECT role FROM match_ratings WHERE fc_id=300").fetchone()[0] == "ALL"


def test_reingest_from_cache(tmp_path):
    from euroleghe_ingest.config import Config
    from euroleghe_ingest.context import Context

    cfg = Config(data_dir=tmp_path / "data", db_path=tmp_path / "data" / "euro.db")
    cfg.cache_dir.mkdir(parents=True)
    (cfg.cache_dir / "ratings_euroleghe_2023-24_md5.xlsx").write_bytes(_xlsx_bytes(ROWS))
    ctx = Context(config=cfg, conn=init_db(cfg.db_path))
    ratings.reingest_from_cache(ctx)
    assert ctx.conn.execute("SELECT COUNT(*) FROM match_ratings").fetchone()[0] == 4
    # season + matchday recovered from the file name
    row = ctx.conn.execute("SELECT DISTINCT season, matchday FROM match_ratings").fetchone()
    assert tuple(row) == ("2023-24", 5)


def test_backfill_clubs_from_ratings(tmp_path):
    from euroleghe_ingest.config import Config
    from euroleghe_ingest.context import Context
    from euroleghe_ingest.modules import rosters

    cfg = Config(data_dir=tmp_path / "data", db_path=tmp_path / "data" / "euro.db")
    (tmp_path / "data").mkdir()
    conn = init_db(cfg.db_path)
    conn.execute("INSERT INTO players(fc_id, canonical_name) VALUES (1, 'X')")
    conn.execute("INSERT INTO rosters(fc_id, season, league) VALUES (1, '2024-25', 'serie_a')")  # no club
    conn.executemany(
        "INSERT INTO match_ratings(fc_id, season, matchday, team, mv) VALUES (?, ?, ?, ?, ?)",
        [(1, "2024-25", 1, "Inter", 6.0), (1, "2024-25", 2, "Inter", 6.5)],
    )
    conn.commit()
    rosters.backfill_clubs(Context(config=cfg, conn=conn))
    row = conn.execute(
        "SELECT c.canonical_name, c.league FROM rosters r JOIN clubs c ON c.fc_club_id = r.fc_club_id "
        "WHERE r.fc_id = 1"
    ).fetchone()
    assert tuple(row) == ("Inter", "serie_a")


def test_backfill_rosters_from_ratings(tmp_path):
    from euroleghe_ingest.config import Config
    from euroleghe_ingest.context import Context
    from euroleghe_ingest.modules import rosters

    cfg = Config(data_dir=tmp_path / "data", db_path=tmp_path / "data" / "euro.db")
    (tmp_path / "data").mkdir()
    conn = init_db(cfg.db_path)
    conn.execute("INSERT INTO players(fc_id, canonical_name) VALUES (7, 'SerieAonly')")
    conn.executemany(
        "INSERT INTO match_ratings(fc_id, season, matchday, role, team, platform, mv) "
        "VALUES (?, ?, ?, ?, ?, ?, ?)",
        [(7, "2023-24", 1, "C", "Cagliari", "default", 6.0),
         (7, "2023-24", 2, "C", "Cagliari", "default", 6.5)],
    )
    conn.commit()
    rosters.backfill_rosters_from_ratings(Context(config=cfg, conn=conn))
    row = conn.execute(
        "SELECT c.canonical_name, r.role_classic, r.league FROM rosters r "
        "JOIN clubs c ON c.fc_club_id = r.fc_club_id WHERE r.fc_id = 7 AND r.season = '2023-24'"
    ).fetchone()
    assert tuple(row) == ("Cagliari", "C", "serie_a")


def test_fix_club_leagues(tmp_path):
    """La lega di un club viene dalle sue PARTITE, mai dalle righe di `rosters`.

    Il test dell'anello. La versione precedente leggeva la maggioranza delle righe di `rosters`, e
    `backfill_rosters_from_ratings` scrive in quelle righe la lega del CLUB: bastava un uomo passato in
    Serie A per tirarci dentro tutta la squadra. Costo misurato il 19/08/2026: otto club stranieri o di
    Serie B archiviati come `serie_a` - Leicester, Everton, Nizza, Valencia, Wolfsburg, Hertha, Palermo,
    Pescara - e 419 righe che entravano nella popolazione di Serie A delle finestre vecchie del gate.

    Qui il caso e' quello vero, in piccolo: il Leicester ha DUE righe di `rosters` che dicono `serie_a`
    (un uomo che poi e' venuto in Italia) e tutte le sue PARTITE in Premier. Vince la Premier.
    """
    from euroleghe_ingest.config import Config
    from euroleghe_ingest.context import Context
    from euroleghe_ingest.modules import rosters

    cfg = Config(data_dir=tmp_path / "data", db_path=tmp_path / "data" / "euro.db")
    (tmp_path / "data").mkdir()
    conn = init_db(cfg.db_path)
    conn.executemany("INSERT INTO clubs(fc_club_id, canonical_name, league) VALUES (?, ?, ?)",
                     [(1, "Leicester City", "serie_a"),      # l'anello lo aveva reso italiano
                      (2, "Hertha Berlino", "serie_a"),      # nessuna partita che la nomini
                      (3, "Napoli", "serie_a")])             # sano, e non si tocca
    conn.executemany("INSERT INTO players(fc_id, canonical_name) VALUES (?, ?)",
                     [(1, "A"), (2, "B"), (3, "C"), (4, "D")])
    conn.executemany("INSERT INTO rosters(fc_id, season, fc_club_id, league) VALUES (?, ?, ?, ?)", [
        (1, "2022-23", 1, "serie_a"), (2, "2022-23", 1, "serie_a"),   # le due righe che mentivano
        (3, "2022-23", 2, "serie_a"), (4, "2022-23", 3, "serie_a"),
    ])
    conn.executemany(
        "INSERT INTO external_match_stats(fc_id, season, source, match_id, competition, club) "
        "VALUES (?, ?, 'sofascore', ?, ?, ?)",
        [(1, "2022-23", f"m{n}", "premier_league", "Leicester City") for n in range(10)]
        + [(4, "2022-23", f"n{n}", "serie_a", "Napoli") for n in range(10)])
    conn.commit()
    rosters.fix_club_leagues(Context(config=cfg, conn=conn))

    league = dict(conn.execute("SELECT canonical_name, league FROM clubs"))
    assert league["Leicester City"] == "premier_league", "le partite battono le righe di rosters"
    assert league["Napoli"] == "serie_a", "un club sano non si tocca"
    assert league["Hertha Berlino"] is None, (
        "dice Serie A e su `default` non ha giocato mai: e' una contraddizione, e ignoto e' meglio di falso")
    # ...e le righe seguono il club, o il filtro `r.league = 'serie_a'` le lascerebbe passare comunque.
    rows = dict(conn.execute(
        "SELECT r.fc_id, r.league FROM rosters r JOIN clubs c ON c.fc_club_id = r.fc_club_id"))
    assert rows[1] == rows[2] == "premier_league"
    assert rows[3] is None
    assert rows[4] == "serie_a"


def test_derive_season_stats_from_ratings(tmp_path):
    from euroleghe_ingest.config import Config
    from euroleghe_ingest.context import Context
    from euroleghe_ingest.modules import stats

    cfg = Config(data_dir=tmp_path / "data", db_path=tmp_path / "data" / "euro.db")
    (tmp_path / "data").mkdir()
    conn = init_db(cfg.db_path)
    conn.executemany("INSERT INTO players(fc_id, canonical_name) VALUES (?, ?)", [(9, "Old"), (8, "Listone")])
    conn.executemany(
        "INSERT INTO match_ratings(fc_id, season, matchday, role, mv, fantavoto, goals, assists) "
        "VALUES (?, ?, ?, ?, ?, ?, ?, ?)",
        [(9, "2016-17", 1, "A", 6.0, 9.0, 1, 0), (9, "2016-17", 2, "A", 7.0, 7.0, 0, 1),
         (8, "2016-17", 1, "C", 5.0, 5.0, 0, 0)],
    )
    # a player already in season_stats (from a listone) must NOT be overwritten
    conn.execute("INSERT INTO season_stats(fc_id, season, pv, mv, fm) VALUES (8, '2016-17', 30, 6.9, 7.7)")
    conn.commit()

    stats.derive_from_ratings(Context(config=cfg, conn=conn))
    derived = conn.execute("SELECT pv, mv, fm, goals, assists FROM season_stats WHERE fc_id=9 AND season='2016-17'").fetchone()
    assert tuple(derived) == (2, 6.5, 8.0, 1, 1)
    kept = conn.execute("SELECT pv, mv, fm FROM season_stats WHERE fc_id=8 AND season='2016-17'").fetchone()
    assert tuple(kept) == (30, 6.9, 7.7)   # listone value untouched


def test_derive_season_stats_per_platform(tmp_path):
    """euro (listone, EuroLeghe subset) and default (full real season) coexist as separate rows,
    so a player's full-season goals count even when they fall outside the EuroLeghe calendar."""
    from euroleghe_ingest.config import Config
    from euroleghe_ingest.context import Context
    from euroleghe_ingest.modules import stats

    cfg = Config(data_dir=tmp_path / "data", db_path=tmp_path / "data" / "euro.db")
    (tmp_path / "data").mkdir()
    conn = init_db(cfg.db_path)
    conn.execute("INSERT INTO players(fc_id, canonical_name) VALUES (5, 'SerieA')")
    # EuroLeghe listone: only the euro-calendar subset -> no goals recorded there
    conn.execute("INSERT INTO season_stats(fc_id, season, platform, pv, mv, fm, goals) "
                 "VALUES (5, '2023-24', 'euro', 2, 6.0, 6.0, 0)")
    # full Serie A voti (default): the player actually scored 2 goals across the season
    conn.executemany(
        "INSERT INTO match_ratings(fc_id, season, matchday, role, platform, mv, fantavoto, goals) "
        "VALUES (?, ?, ?, ?, ?, ?, ?, ?)",
        [(5, "2023-24", 1, "D", "default", 6.0, 9.0, 1),
         (5, "2023-24", 2, "D", "default", 6.0, 6.0, 0),
         (5, "2023-24", 3, "D", "default", 7.0, 10.0, 1)],
    )
    conn.commit()

    stats.derive_from_ratings(Context(config=cfg, conn=conn))
    euro = conn.execute("SELECT goals FROM season_stats WHERE fc_id=5 AND platform='euro'").fetchone()
    assert euro[0] == 0                                    # target perspective untouched
    default = conn.execute("SELECT pv, goals FROM season_stats WHERE fc_id=5 AND platform='default'").fetchone()
    assert tuple(default) == (3, 2)                        # full-season propensity captured


def test_parse_and_upsert_listone(tmp_path):
    """The listone (quotazioni) fills Mantra roles (RM) + price on rosters, incl. non-top teams
    whose roster came from the voti (Classic role only)."""
    header = ["Id", "R", "RM", "Nome", "Squadra", "Qt.A", "Qt.I", "Diff.",
              "Qt.A M", "Qt.I M", "Diff.M", "FVM", "FVM M"]
    rows = [
        ["Quotazioni Fantacalcio Stagione 2023 24"] + [None] * 12,
        header,
        [111, "P", "Por", "Tizio", "Cagliari", 5, 5, 0, 5, 5, 0, 10, 10],
        [222, "D", "Dc;Ds", "Caio", "Torino", 12, 10, 2, 14, 12, 2, 40, 45],
    ]
    wb = Workbook()
    ws = wb.active
    ws.title = "Tutti"
    for r in rows:
        ws.append(r)
    ced = wb.create_sheet("Ceduti")           # a mid-season departure: still played -> in the voti
    ced.append(["Quotazioni Fantacalcio Stagione 2023 24 - Ceduti"] + [None] * 12)
    ced.append(header)
    ced.append([333, "D", "Dd;Dc", "Sold", "Salernitana", 2, 3, -1, 2, 3, -1, 5, 6])
    buf = io.BytesIO()
    wb.save(buf)
    data = buf.getvalue()

    recs = ratings.parse_listone(data, "2023-24")
    assert len(recs) == 3
    by = {r["fc_id"]: r for r in recs}
    assert by[222]["roles"] == ["dc", "ds"] and by[222]["role_classic"] == "D"
    assert by[111]["team"] == "Cagliari" and by[111]["price"] == 5
    assert by[333]["roles"] == ["dd", "dc"] and by[333]["team"] == "Salernitana"   # from 'Ceduti'

    conn = init_db(tmp_path / "euro.db")
    conn.execute("INSERT INTO players(fc_id, canonical_name) VALUES (222, 'Caio')")
    # a roster reconstructed from the voti: Classic role only, no Mantra role
    conn.execute("INSERT INTO rosters(fc_id, season, role_classic, league) VALUES (222,'2023-24','D','serie_a')")
    conn.commit()

    n = ratings.upsert_listone(conn, "2023-24", recs, "default")
    assert n == 3
    filled = conn.execute("SELECT roles, price FROM rosters WHERE fc_id=222 AND season='2023-24'").fetchone()
    assert filled[0] == "dc;ds" and filled[1] == 12         # Mantra roles + price now present
    created = conn.execute(
        "SELECT c.canonical_name, r.league, r.roles FROM rosters r JOIN clubs c "
        "ON c.fc_club_id = r.fc_club_id WHERE r.fc_id = 111 AND r.season='2023-24'").fetchone()
    assert tuple(created) == ("Cagliari", "serie_a", "por")  # brand-new player got a full roster row


def test_ratings_consistency_check(tmp_path):
    conn = init_db(tmp_path / "euro.db")
    for fc_id in (100, 200):
        conn.execute("INSERT INTO players(fc_id, canonical_name) VALUES (?, ?)", (fc_id, f"P{fc_id}"))
    # season aggregates
    conn.execute("INSERT INTO season_stats(fc_id, season, pv, mv, fm) VALUES (100,'2023-24',2,6.0,5.0)")
    conn.execute("INSERT INTO season_stats(fc_id, season, pv, mv, fm) VALUES (200,'2023-24',1,7.0,7.0)")
    # fc 100: two matchdays that average to Mv=6, FM=5 -> consistent
    conn.execute("INSERT INTO match_ratings(fc_id, season, matchday, mv, fantavoto) VALUES (100,'2023-24',1,6.0,5.0)")
    conn.execute("INSERT INTO match_ratings(fc_id, season, matchday, mv, fantavoto) VALUES (100,'2023-24',2,6.0,5.0)")
    # fc 200: one matchday with Mv=8 but season says 7 -> Mv mismatch (hard)
    conn.execute("INSERT INTO match_ratings(fc_id, season, matchday, mv, fantavoto) VALUES (200,'2023-24',1,8.0,8.0)")
    conn.commit()

    problems = check_ratings_consistency(conn)
    assert any("fc_id=200" in p for p in problems)
    assert not any("fc_id=100" in p for p in problems)

def test_a_listone_says_which_season_it_is_and_a_new_one_is_found_before_the_first_matchday():
    """Every August the listone is published weeks before a single vote exists, and the championship id
    used to be readable only off the VOTI page - which for a season with no matchdays carries none. So the
    quotazioni page answers instead (Serie A 2026-27 reads 21 there, 2025-26 read 20).

    It is a fallback and it needs a guard, because those pages serve «the current list» whatever season is
    asked of them: the euro one still reads 108 = 2025-26. The guard is the workbook's own title cell, so a
    file that does not say the season being ingested is refused instead of filed under the wrong year.
    """
    wb = Workbook()
    ws = wb.active
    ws.title = "Tutti"
    ws.append(["Quotazioni Fantacalcio Stagione 2026 27"])
    ws.append(["Id", "R", "RM", "Nome", "Squadra", "Qt.A", "Qt.I", "Diff."])
    ws.append([5841, "P", "Por", "Svilar", "Roma", 18, 18, 0])
    buf = io.BytesIO()
    wb.save(buf)
    assert ratings.listone_season(buf.getvalue()) == "2026-27"

    euro = Workbook()
    sheet = euro.active
    sheet.title = "Tutti"
    sheet.append(["Quotazioni Fantacalcio EuroLeghe Stagione 2025 26"])
    sheet.append(["Id", "R", "RM", "Nome", "Nazione", "Squadra", "Qt.A", "Qt.I"])
    other = io.BytesIO()
    euro.save(other)
    assert ratings.listone_season(other.getvalue()) == "2025-26", "and it reads the euro title too"

    empty = Workbook()
    empty.active.append(["no title here"])
    nothing = io.BytesIO()
    empty.save(nothing)
    assert ratings.listone_season(nothing.getvalue()) is None, "no claim is not a wrong claim"

    # the id is read from the QUOTAZIONI page when the voti page has none, and from the voti page first
    class Reply:
        def __init__(self, text):
            self.status_code, self.text = 200, text

    pages = {}

    class Session:
        pass

    def fake_http(_session, _method, url, **_kwargs):
        return Reply(pages.get(url, ""))

    original = ratings._http
    ratings._http = fake_http
    try:
        pages = {ratings.PLATFORMS["default"].format(season="2025-26"):
                 '<a href="/api/v1/Excel/votes/20/1">',
                 ratings.PRICE_PAGES["default"].format(season="2026-27"):
                 '<a href="/api/v1/Excel/prices/21/1">'}
        assert ratings.resolve_championship_id(Session(), "default", "2025-26") == "20"
        assert ratings.resolve_championship_id(Session(), "default", "2026-27") == "21"
        assert ratings.resolve_championship_id(Session(), "default", "2027-28") is None
    finally:
        ratings._http = original
