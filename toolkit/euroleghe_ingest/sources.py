"""Unified reading of the raw datasets (CSV 23/24-24/25 + Excel 25/26).

Exposes `iter_records(config)`, which yields NORMALIZED records, one per (player, season),
regardless of the source format. rosters/stats consume these records.

Known differences between sources (from probing the real files):
- CSV (2023-24, 2024-25): 17 columns, lowercase; penalties as rig_s (scored) / rig_t (taken) /
  rp (saved); NO own goals or nationality; accents in names lost at the source (U+FFFD);
  the `squadra` (club) column is entirely empty in 2024-25 -> club unknown for that season.
- Excel 2025-26: header on row 2, sheet 'Tutti'; penalties as R+ (scored) / R- (missed) /
  Rc (taken) / Rp (saved); Nazione and Au (own goals) present; Rm is multi-role with ';'.
  NOTE: in the Excel the column labeled 'Nazione' actually holds the LEAGUE, not nationality.
"""

from __future__ import annotations

import csv
import re
from dataclasses import dataclass
from pathlib import Path

from openpyxl import load_workbook

from euroleghe_ingest.config import Config

# (season, file name in data/raw, format)
SEASON_SOURCES: tuple[tuple[str, str, str], ...] = (
    ("2023-24", "euroleghe-stats-2023-24.csv", "csv"),
    ("2024-25", "euroleghe-stats-2024-25.csv", "csv"),
    ("2025-26", "Statistiche_Fantacalcio_EuroLeghe_Stagione_2025_26.xlsx", "xlsx"),
)

LEAGUE_MAP: dict[str, str] = {
    "serie a": "serie_a",
    "premier league": "premier_league",
    "liga": "la_liga",
    "liga1": "ligue_1",
    "bundesliga": "bundesliga",
}

# Canonical role vocabulary. Ingestion normalizes to these (Classic uppercase, Mantra lowercase);
# `validate` surfaces anything outside these sets instead of silently accepting it.
CLASSIC_ROLES: tuple[str, ...] = ("P", "D", "C", "A")
MANTRA_ROLES: tuple[str, ...] = ("por", "dc", "dd", "ds", "b", "e", "m", "c", "w", "t", "a", "pc")

# Which Mantra roles belong to each Classic role. A Mantra squad is still rostered by Classic role
# (3/8/8/6 on fantacalcio.it), so this is what lets a Classic squad rule constrain Mantra role depth.
MANTRA_BY_CLASSIC: dict[str, tuple[str, ...]] = {
    "P": ("por",),
    "D": ("dc", "dd", "ds", "b"),
    "C": ("e", "m", "c"),
    "A": ("w", "t", "a", "pc"),
}


@dataclass
class Record:
    """Normalized row for a (player, season)."""

    fc_id: int
    season: str
    name: str
    nationality: str | None
    league: str | None
    club: str | None
    role_classic: str | None
    roles: list[str]
    pv: int | None
    mv: float | None
    fm: float | None
    goals: int | None
    assists: int | None
    yellows: int | None
    reds: int | None
    own_goals: int | None
    pen_scored: int | None
    pen_missed: int | None
    goals_conceded: int | None
    pen_saved: int | None


# ---------- conversion helpers ----------
def _to_int(value) -> int | None:
    if value is None or value == "":
        return None
    try:
        return int(float(value))
    except (TypeError, ValueError):
        return None


def _to_float(value) -> float | None:
    if value is None or value == "":
        return None
    try:
        return float(value)
    except (TypeError, ValueError):
        return None


def _norm_league(value: str | None) -> str | None:
    if not value:
        return None
    return LEAGUE_MAP.get(value.strip().lower())


def _norm_roles(value: str | None) -> list[str]:
    if not value:
        return []
    # Mantra roles can be multi-valued; the CSVs separate them with '|', the Excel with ';'.
    return [r.strip().lower() for r in re.split(r"[;/|]", str(value)) if r.strip()]


def _clean(value) -> str | None:
    if value is None:
        return None
    text = str(value).strip()
    return text or None


# ---------- per-format readers ----------
def _read_csv(path: Path, season: str):
    # utf-8-sig strips a UTF-8 BOM; otherwise the first header becomes "﻿id" and every row is
    # dropped (fc_id None), silently losing the whole season.
    text = path.read_bytes().decode("utf-8-sig", errors="replace")
    reader = csv.DictReader(text.splitlines())
    for row in reader:
        fc_id = _to_int(row.get("id"))
        if fc_id is None:
            continue
        scored = _to_int(row.get("rig_s"))
        taken = _to_int(row.get("rig_t"))
        missed = (taken - scored) if (scored is not None and taken is not None) else None
        yield Record(
            fc_id=fc_id,
            season=season,
            name=_clean(row.get("nome")) or str(fc_id),
            nationality=None,
            league=_norm_league(row.get("lega")),
            club=_clean(row.get("squadra")),
            role_classic=_clean(row.get("r")),
            roles=_norm_roles(row.get("rm")),
            pv=_to_int(row.get("pv")),
            mv=_to_float(row.get("mv")),
            fm=_to_float(row.get("fm")),
            goals=_to_int(row.get("gf")),
            assists=_to_int(row.get("ass")),
            yellows=_to_int(row.get("amm")),
            reds=_to_int(row.get("esp")),
            own_goals=None,  # not present in the CSVs
            pen_scored=scored,
            pen_missed=missed,
            goals_conceded=_to_int(row.get("gs")),
            pen_saved=_to_int(row.get("rp")),
        )


def _read_xlsx(path: Path, season: str):
    wb = load_workbook(path, read_only=True, data_only=True)
    try:
        ws = wb["Tutti"] if "Tutti" in wb.sheetnames else wb[wb.sheetnames[0]]
        rows = ws.iter_rows(min_row=2, values_only=True)  # row 2 = header
        header = [str(c).strip() if c is not None else "" for c in next(rows)]
        idx = {name: i for i, name in enumerate(header)}

        def cell(row, key):
            i = idx.get(key)
            return row[i] if i is not None and i < len(row) else None

        for row in rows:
            fc_id = _to_int(cell(row, "Id"))
            if fc_id is None:
                continue
            yield Record(
                fc_id=fc_id,
                season=season,
                name=_clean(cell(row, "Nome")) or str(fc_id),
                # NOTE: in the Excel the 'Nazione' column actually holds the LEAGUE
                # (e.g. 'Liga', 'Serie A'), not nationality. No source provides nationality.
                nationality=None,
                league=_norm_league(cell(row, "Nazione")),
                club=_clean(cell(row, "Squadra")),
                role_classic=_clean(cell(row, "R")),
                roles=_norm_roles(cell(row, "Rm")),
                pv=_to_int(cell(row, "Pv")),
                mv=_to_float(cell(row, "Mv")),
                fm=_to_float(cell(row, "Fm")),
                goals=_to_int(cell(row, "Gf")),
                assists=_to_int(cell(row, "Ass")),
                yellows=_to_int(cell(row, "Amm")),
                reds=_to_int(cell(row, "Esp")),
                own_goals=_to_int(cell(row, "Au")),
                pen_scored=_to_int(cell(row, "R+")),
                pen_missed=_to_int(cell(row, "R-")),
                goals_conceded=_to_int(cell(row, "Gs")),
                pen_saved=_to_int(cell(row, "Rp")),
            )
    finally:
        wb.close()


def available_sources(config: Config) -> list[tuple[str, Path, str]]:
    """Sources actually present in data/raw."""
    out = []
    for season, filename, fmt in SEASON_SOURCES:
        path = config.raw_dir / filename
        if path.exists():
            out.append((season, path, fmt))
    return out


def iter_records(config: Config):
    """Yield normalized Records from every available source."""
    for season, path, fmt in available_sources(config):
        reader = _read_csv if fmt == "csv" else _read_xlsx
        yield from reader(path, season)
