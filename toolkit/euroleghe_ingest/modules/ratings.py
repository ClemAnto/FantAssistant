"""ratings - per-matchday player ratings from the OFFICIAL fantacalcio.it Excel.

Auth: logs in (POST /api/v1/User/login with credentials from .env) to obtain the session cookie,
then downloads the official Excel per matchday (GET /api/v1/Excel/votes/{championshipId}/{matchday}),
which is the clean ground truth. The public HTML voti page is deliberately NOT used: it is
boobytrapped for scrapers (values only in data-value, a "55" poison on the grade of players who did
not play, ~97% of rows hidden). The authenticated Excel has none of that.

Aggregation (option A): canonical, cross-season-comparable fields go into match_ratings; every raw
bonus column of the source Excel is also stored as-is in match_rating_bonuses, so season-specific
bonuses (assist subtypes, player-of-the-match, ...) are never lost when the site changes them.

Coaches (Ruolo="ALL") are kept too, for info, tagged via match_ratings.role.

Polite scraping: a delay with jitter between requests to avoid tripping defenses.
"""

from __future__ import annotations

import io
import os
import random
import re
import time

import requests
from dotenv import load_dotenv
from openpyxl import load_workbook

from euroleghe_ingest.context import Context
from euroleghe_ingest.sources import _norm_roles

NAME = "ratings"
DESCRIPTION = "Official per-matchday Excel -> match_ratings (+ raw bonuses)"
DEPENDS_ON: list[str] = ["rosters"]
RAW_INPUTS: list[str] = []
NETWORK = True

BASE_URL = "https://www.fantacalcio.it"
LOGIN_ENDPOINT = BASE_URL + "/api/v1/User/login"
EXCEL_ENDPOINT = BASE_URL + "/api/v1/Excel/votes/{cid}/{matchday}"
# Listone (quotazioni): the season's Mantra roles (RM) + prices for ALL teams. Same championship id
# as the votes (verified: default 18/19/20, euro 106/107/108), listType 1 = the full list.
PRICES_ENDPOINT = BASE_URL + "/api/v1/Excel/prices/{cid}/1"
_EXCEL_HREF = re.compile(r"/api/v1/Excel/votes/(\d+)/")
_CACHE_NAME = re.compile(r"ratings_(?:([a-z_]+)_)?(\d{4}-\d{2})_md(\d+)\.xlsx$")
_LISTONE_NAME = re.compile(r"listone_([a-z_]+)_(\d{4}-\d{2})\.xlsx$")

# Platforms and their voti-page URL (the championship id + Excel link are read from that page).
# 'euro' = EuroLeghe (5 leagues, top clubs only - Serie A is PARTIAL); 'default' = the classic
# fantacalcio.it Serie A (all 20 teams). NOTE: the two use DIFFERENT matchday calendars.
PLATFORMS: dict[str, str] = {
    "euro": BASE_URL + "/voti-fantacalcio-euro-leghe/{season}/1",
    "default": BASE_URL + "/voti-fantacalcio-serie-a/{season}/1",
}
DEFAULT_PLATFORM = "euro"

# Old cache files used the pre-rename tokens; map them to the platform values.
_PLATFORM_ALIAS = {"euroleghe": "euro", "serie_a": "default"}
SEASONS: tuple[str, ...] = ("2023-24", "2024-25", "2025-26")   # default when no season is selected
MAX_MATCHDAYS = 60

# Polite rate limiting (seconds): base + uniform jitter between every request.
REQUEST_DELAY = 3.0
REQUEST_JITTER = 2.0

# Source Excel header (Italian abbreviations) -> canonical match_ratings column.
CANON: dict[str, str] = {
    "Voto": "mv", "Gf": "goals", "Gs": "goals_conceded", "Rp": "pen_saved",
    "Rs": "pen_scored", "Rf": "pen_missed", "Au": "own_goals",
    "Amm": "yellows", "Esp": "reds", "Ass": "assists",
}
_META_COLS = {"Cod.", "Ruolo", "Squadra", "Nome"}
_PLAYER_ROLES = {"P", "D", "C", "A"}


# ---------- helpers ----------
def _num(v):
    if v is None or v == "":
        return None
    try:
        return float(v)
    except (TypeError, ValueError):
        return None


def _polite_sleep(cancel_event=None) -> None:
    delay = REQUEST_DELAY + random.uniform(0, REQUEST_JITTER)  # noqa: S311 - not crypto
    if cancel_event is not None:
        cancel_event.wait(delay)   # wakes up immediately if cancellation is requested
    else:
        time.sleep(delay)


def _credentials(config) -> tuple[str, str]:
    load_dotenv(config.repo_root / ".env")
    user = os.environ.get("FANTACALCIO_USERNAME")
    pwd = os.environ.get("FANTACALCIO_PASSWORD")
    if not user or not pwd:
        raise RuntimeError("Missing FANTACALCIO_USERNAME / FANTACALCIO_PASSWORD in .env")
    return user, pwd


def _new_session(user_agent: str) -> requests.Session:
    s = requests.Session()
    s.headers.update({"User-Agent": user_agent, "Content-Type": "application/json"})
    return s


# ---------- network ----------
def login(session: requests.Session, username: str, password: str) -> None:
    """POST credentials; the session cookie is stored on the session on success."""
    r = session.post(LOGIN_ENDPOINT, json={"username": username, "password": password}, timeout=30)
    try:
        data = r.json()
    except ValueError:
        raise RuntimeError(f"Login: unexpected response (HTTP {r.status_code})")
    if not data.get("success"):
        msgs = "; ".join(e.get("message", "") for e in data.get("errors", [])) or f"HTTP {r.status_code}"
        raise RuntimeError(f"Login failed: {msgs}")   # never log the password


def resolve_championship_id(session: requests.Session, platform: str, season: str) -> str | None:
    r = session.get(PLATFORMS[platform].format(season=season), timeout=30)
    if r.status_code != 200:
        return None
    m = _EXCEL_HREF.search(r.text)
    return m.group(1) if m else None


def download_matchday(session: requests.Session, cid: str, matchday: int) -> bytes | None:
    r = session.get(EXCEL_ENDPOINT.format(cid=cid, matchday=matchday), timeout=30)
    if r.status_code != 200:
        return None
    content = r.content
    return content if content[:2] == b"PK" else None   # xlsx = zip (starts with PK)


def download_listone(session: requests.Session, cid: str) -> bytes | None:
    """Download the season listone (quotazioni) Excel for a championship id (Mantra roles + prices)."""
    r = session.get(PRICES_ENDPOINT.format(cid=cid), timeout=30)
    if r.status_code != 200:
        return None
    return r.content if r.content[:2] == b"PK" else None


# ---------- parsing (pure, offline-testable) ----------
def parse_workbook(data: bytes, season: str, matchday: int) -> list[dict]:
    """Parse one official votes Excel into per-player records (players + coaches).

    Handles both layouts: EuroLeghe has a 'Squadra' column; the classic Serie A has none and instead
    prints the club as a separator row before each team block (with the header repeated per team)."""
    wb = load_workbook(io.BytesIO(data), read_only=True, data_only=True)
    try:
        ws = wb["Statistico"] if "Statistico" in wb.sheetnames else wb[wb.sheetnames[0]]
        header_idx: dict[str, int] = {}
        current_team: str | None = None
        records: list[dict] = []
        for row in ws.iter_rows(values_only=True):
            first = row[0]
            if isinstance(first, str):
                s = first.strip()
                if s.startswith("Cod"):
                    header_idx = {str(c).strip(): i for i, c in enumerate(row) if c is not None}
                elif all(c is None for c in row[1:]) and len(s) <= 30 and "fantacalcio" not in s.lower():
                    current_team = s          # team-separator row (not a disclaimer)
                continue
            if not isinstance(first, int) or not header_idx:
                continue
            team = None
            if "Squadra" in header_idx:
                v = row[header_idx["Squadra"]]
                team = str(v).strip() if v is not None else None
            rec = {
                "fc_id": first,
                "season": season,
                "matchday": matchday,
                "role": (str(row[header_idx["Ruolo"]]).strip() if "Ruolo" in header_idx else None),
                "name": (str(row[header_idx["Nome"]]).strip() if "Nome" in header_idx else None),
                "team": team or current_team,
                "canon": {},
                "raw": {},
            }
            for head, i in header_idx.items():
                if head in _META_COLS or head == "":
                    continue
                value = row[i] if i < len(row) else None
                rec["raw"][head] = _num(value)
                canon = CANON.get(head) or ("assists" if head.lower().startswith("ass") else None)
                if canon == "assists":
                    rec["canon"]["assists"] = (rec["canon"].get("assists") or 0) + (_num(value) or 0)
                elif canon:
                    rec["canon"][canon] = _num(value)
            records.append(rec)
        return records
    finally:
        wb.close()


def parse_listone(data: bytes, season: str) -> list[dict]:
    """Parse a listone (quotazioni) Excel into per-player role/price records (sheet 'Tutti').

    Header: Id | R | RM | Nome | Squadra | Qt.A | Qt.I | ... (Id=fc_id, R=Classic role,
    RM=Mantra roles, Qt.A=current auction price). Rows above the header (the season title) and any
    non-integer Id are skipped."""
    wb = load_workbook(io.BytesIO(data), read_only=True, data_only=True)
    try:
        ws = wb["Tutti"] if "Tutti" in wb.sheetnames else wb[wb.sheetnames[0]]
        header: dict[str, int] = {}
        out: list[dict] = []
        for row in ws.iter_rows(values_only=True):
            first = row[0]
            if isinstance(first, str) and first.strip() == "Id":
                header = {str(c).strip(): i for i, c in enumerate(row) if c is not None}
                continue
            if not isinstance(first, int) or not header:
                continue

            def cell(col):
                i = header.get(col)
                return row[i] if (i is not None and i < len(row)) else None

            out.append({
                "fc_id": first,
                "season": season,
                "role_classic": (str(cell("R")).strip() if cell("R") is not None else None),
                "roles": _norm_roles(cell("RM")),
                "name": (str(cell("Nome")).strip() if cell("Nome") is not None else None),
                "team": (str(cell("Squadra")).strip() if cell("Squadra") is not None else None),
                "price": _num(cell("Qt.A")),
            })
        return out
    finally:
        wb.close()


def compute_fantavoto(canon: dict, scoring: dict[str, float]) -> float | None:
    mv = canon.get("mv")
    if mv is None:
        return None
    val = (
        mv
        + scoring["goal_bonus"] * (canon.get("goals") or 0)
        + scoring["assist_bonus"] * (canon.get("assists") or 0)
        - scoring["own_goal_malus"] * (canon.get("own_goals") or 0)
        + scoring["penalty_scored_bonus"] * (canon.get("pen_scored") or 0)
        - scoring["penalty_missed_malus"] * (canon.get("pen_missed") or 0)
        + scoring["penalty_saved_bonus_gk"] * (canon.get("pen_saved") or 0)
        - scoring["goal_conceded_malus_gk"] * (canon.get("goals_conceded") or 0)
        - scoring["yellow_card_malus"] * (canon.get("yellows") or 0)
        - scoring["red_card_malus"] * (canon.get("reds") or 0)
    )
    return round(val, 2)


# ---------- persistence ----------
def upsert_records(conn, records: list[dict], scoring: dict[str, float],
                   platform: str = DEFAULT_PLATFORM) -> int:
    for rec in records:
        conn.execute(
            "INSERT OR IGNORE INTO players(fc_id, canonical_name) VALUES (?, ?)",
            (rec["fc_id"], rec["name"] or str(rec["fc_id"])),
        )
        c = rec["canon"]
        mv = c.get("mv")
        if mv is not None and not (0 <= mv <= 10):   # guard against poisoned values
            mv = None
        c = {**c, "mv": mv}
        status = "played" if mv is not None else "no_vote"
        conn.execute(
            """
            INSERT OR REPLACE INTO match_ratings(
                fc_id, season, matchday, role, team, platform, mv, goals, assists, own_goals,
                pen_scored, pen_missed, pen_saved, goals_conceded, yellows, reds,
                fantavoto, status)
            VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?)
            """,
            (
                rec["fc_id"], rec["season"], rec["matchday"], rec["role"], rec.get("team"), platform, mv,
                c.get("goals"), c.get("assists"), c.get("own_goals"), c.get("pen_scored"),
                c.get("pen_missed"), c.get("pen_saved"), c.get("goals_conceded"),
                c.get("yellows"), c.get("reds"), compute_fantavoto(c, scoring), status,
            ),
        )
        for key, value in rec["raw"].items():
            conn.execute(
                "INSERT OR REPLACE INTO match_rating_bonuses"
                "(fc_id, season, matchday, platform, bonus_key, value) VALUES (?, ?, ?, ?, ?, ?)",
                (rec["fc_id"], rec["season"], rec["matchday"], platform, key, value),
            )
    return len(records)


def upsert_listone(conn, season: str, records: list[dict], platform: str = DEFAULT_PLATFORM) -> int:
    """Enrich rosters with the listone: Mantra roles (RM), Classic role, price and club. This fills
    the non-top Serie A players whose rosters came from the voti (Classic role only, no Mantra).
    The listone is authoritative for roles/role/price; the roster league is preserved when known."""
    from euroleghe_ingest.modules.rosters import _get_or_create_club

    n = 0
    for rec in records:
        conn.execute("INSERT OR IGNORE INTO players(fc_id, canonical_name) VALUES (?, ?)",
                     (rec["fc_id"], rec["name"] or str(rec["fc_id"])))
        if platform == "default":
            league = "serie_a"
        else:  # euro spans 5 leagues: take the club's known league if we have it, else leave unknown
            lk = conn.execute("SELECT league FROM clubs WHERE canonical_name = ? AND league IS NOT NULL "
                              "LIMIT 1", (rec["team"],)).fetchone()
            league = lk[0] if lk else None
        club_id = _get_or_create_club(conn, rec["team"], league)
        roles = ";".join(rec["roles"]) or None
        conn.execute(
            """
            INSERT INTO rosters(fc_id, season, fc_club_id, roles, role_classic, league, price)
            VALUES (?, ?, ?, ?, ?, ?, ?)
            ON CONFLICT(fc_id, season) DO UPDATE SET
                roles        = COALESCE(excluded.roles, rosters.roles),
                role_classic = COALESCE(excluded.role_classic, rosters.role_classic),
                price        = COALESCE(excluded.price, rosters.price),
                fc_club_id   = COALESCE(excluded.fc_club_id, rosters.fc_club_id),
                league       = COALESCE(rosters.league, excluded.league)
            """,
            (rec["fc_id"], season, club_id, roles, rec["role_classic"], league, rec["price"]),
        )
        n += 1
    return n


# ---------- orchestration ----------
def run(ctx: Context, *, platform: str = DEFAULT_PLATFORM, seasons=None,
        refresh: bool = False, **kwargs) -> None:
    """Scrape ratings for a platform ('euro' or 'default') and one or more seasons.
    Resumable (skips matchdays already in the DB unless refresh=True) and interruptible (Ctrl-C or
    ctx.cancel_event): each matchday is committed as it lands, so a stop never loses what was downloaded."""
    if platform not in PLATFORMS:
        raise RuntimeError(f"Unknown platform {platform!r}; choose from {sorted(PLATFORMS)}")
    if isinstance(seasons, str):
        seasons = [seasons]
    seasons = tuple(seasons) if seasons else SEASONS

    conn = ctx.require_conn()
    user, pwd = _credentials(ctx.config)
    user_agent = os.environ.get("EUROLEGHE_USER_AGENT", "FantAssistant/0.1 (+personal-use)")
    scoring = ctx.config.load_scoring()   # default fantacalcio.it scoring for the stored fantavoto

    session = _new_session(user_agent)
    print(f"[ratings] platform={platform} seasons={list(seasons)} - logging in...")
    login(session, user, pwd)

    ctx.config.cache_dir.mkdir(parents=True, exist_ok=True)
    total = 0
    try:
        for season in seasons:
            if ctx.cancelled():
                break
            cid = resolve_championship_id(session, platform, season)
            if not cid:
                print(f"[ratings] {season}: championship id not found - skipping")
                continue
            # resume PER COMPETITION: EuroLeghe and Serie A share the (season, matchday) key but
            # cover different teams, so a matchday scraped for one is NOT done for the other.
            done = {md for (md,) in conn.execute(
                "SELECT DISTINCT matchday FROM match_ratings WHERE season = ? AND platform = ?",
                (season, platform))}
            note = f" (resuming, {len(done)} matchdays already present)" if done and not refresh else ""
            print(f"[ratings] {season}: championship {cid}{note} - downloading...")
            # Listone (quotazioni), same championship id as the votes: Mantra roles (RM) + prices for
            # ALL teams -> enrich rosters (fills non-top Serie A players who had only the Classic role).
            _polite_sleep(ctx.cancel_event)
            if not ctx.cancelled():
                listone = download_listone(session, cid)
                if listone:
                    (ctx.config.cache_dir / f"listone_{platform}_{season}.xlsx").write_bytes(listone)
                    n_lst = upsert_listone(conn, season, parse_listone(listone, season), platform)
                    conn.commit()
                    print(f"[ratings] {season}: listone {platform} -> {n_lst} players (Mantra roles + prices)")
                else:
                    print(f"[ratings] {season}: listone not available (prices/{cid})")
            season_rows = last_md = 0
            for matchday in range(1, MAX_MATCHDAYS + 1):
                if ctx.cancelled():
                    break
                if not refresh and matchday in done:
                    last_md = matchday
                    continue
                _polite_sleep(ctx.cancel_event)
                if ctx.cancelled():
                    break
                data = download_matchday(session, cid, matchday)
                if data is None:
                    print(f"[ratings] {season}: stop at md{matchday} (HTTP not 200 or not an xlsx)")
                    break
                records = parse_workbook(data, season, matchday)
                if not records:
                    # beyond the last played matchday the endpoint still returns a valid but empty sheet
                    print(f"[ratings] {season}: stop at md{matchday} (empty sheet - end of season)")
                    break
                (ctx.config.cache_dir / f"ratings_{platform}_{season}_md{matchday}.xlsx").write_bytes(data)
                teams = len({r["team"] for r in records if r.get("team")})
                players = sum(1 for r in records if r["role"] in _PLAYER_ROLES)
                voted = sum(1 for r in records if r["canon"].get("mv") is not None)
                total += upsert_records(conn, records, scoring, platform)
                conn.commit()
                season_rows += len(records)
                last_md = matchday
                print(f"[ratings] {season} md{matchday}: {teams} teams, {players} players, {voted} voted")
            print(f"[ratings] {season}: reached md{last_md}, +{season_rows} new rows")
            if ctx.cancelled():
                break
    except KeyboardInterrupt:
        print("[ratings] interrupted - already-downloaded matchdays are saved")
        return
    if ctx.cancelled():
        print("[ratings] stopped by request - already-downloaded matchdays are saved")
        return
    print(f"[ratings] ALL done - {total} new rows across {len(seasons)} seasons")


def reingest_from_cache(ctx: Context) -> None:
    """Rebuild match_ratings offline from the cached Excel files (no network), so a rebuild keeps
    the scraped ratings without re-downloading. The cached xlsx are the raw source of truth."""
    conn = ctx.require_conn()
    scoring = ctx.config.load_scoring()
    files = sorted(ctx.config.cache_dir.glob("ratings_*.xlsx"))
    total = 0
    for path in files:
        m = _CACHE_NAME.search(path.name)
        if not m:
            continue
        token = m.group(1)
        platform = _PLATFORM_ALIAS.get(token, token) if token else DEFAULT_PLATFORM
        season, matchday = m.group(2), int(m.group(3))
        total += upsert_records(conn, parse_workbook(path.read_bytes(), season, matchday),
                                scoring, platform)
    conn.commit()
    if files:
        print(f"[ratings] reingested {total} rows from {len(files)} cached files")


def reingest_listone_from_cache(ctx: Context) -> None:
    """Re-apply the cached listone files to rosters offline (rebuild path), so the Mantra roles +
    prices survive a rebuild without re-downloading."""
    conn = ctx.require_conn()
    files = sorted(ctx.config.cache_dir.glob("listone_*.xlsx"))
    total = 0
    for path in files:
        m = _LISTONE_NAME.search(path.name)
        if not m:
            continue
        token, season = m.group(1), m.group(2)
        platform = _PLATFORM_ALIAS.get(token, token)
        total += upsert_listone(conn, season, parse_listone(path.read_bytes(), season), platform)
    conn.commit()
    if files:
        print(f"[ratings] reingested {total} listone rows from {len(files)} cached files")
